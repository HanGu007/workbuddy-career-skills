#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
求职档案模板数据清除脚本
功能：将模板中的示例数据清空，生成干净的求职档案文件供新用户使用
"""

import openpyxl
import shutil
import os
import sys
from datetime import datetime

def clear_template_data(template_path, output_path=None):
    """
    清除模板中的示例数据，保留表头、格式和公式结构
    """
    if not os.path.exists(template_path):
        print(f"错误：模板文件不存在 {template_path}")
        sys.exit(1)

    # 加载工作簿
    wb = openpyxl.load_workbook(template_path)

    # ========== Sheet 1: 投递总览 ==========
    if '投递总览' in wb.sheetnames:
        ws = wb['投递总览']
        # 清除个人信息（B2-B4）
        ws['B2'].value = None  # 目标岗位
        ws['B3'].value = None  # 目标城市
        ws['B4'].value = None  # 期望薪资
        # 清除统计数据（B7-B12）
        for row in range(7, 13):
            ws.cell(row=row, column=2).value = 0
            ws.cell(row=row, column=3).value = '-'
        # 清除状态分布数据（B17-B22左右）
        for row in range(17, 30):
            cell_b = ws.cell(row=row, column=2)
            cell_c = ws.cell(row=row, column=3)
            if cell_b.value is not None and str(cell_b.value).replace('.', '').isdigit():
                cell_b.value = 0
            if cell_c.value is not None:
                cell_c.value = ''

    # ========== Sheet 2: 投递明细表（综合版）==========
    if '投递明细表（综合版）' in wb.sheetnames:
        ws = wb['投递明细表（综合版）']
        # 保留第1行表头，从第2行开始清除
        max_row = ws.max_row
        for row in range(2, max_row + 1):
            for col in range(1, 18):  # A-Q列
                cell = ws.cell(row=row, column=col)
                # 保留序号列的数字格式，清空内容
                if col == 1:  # 序号列
                    cell.value = None
                else:
                    cell.value = None

    # ========== Sheet 3: 面试准备清单 ==========
    if '面试准备清单' in wb.sheetnames:
        ws = wb['面试准备清单']
        max_row = ws.max_row
        for row in range(2, max_row + 1):
            for col in range(1, 8):  # A-G列
                ws.cell(row=row, column=col).value = None

    # ========== Sheet 4: 薪资对比分析 ==========
    if '薪资对比分析' in wb.sheetnames:
        ws = wb['薪资对比分析']
        # 清空第1行的公司名称（B1-D1），保留A1"对比项"
        for col in range(2, 5):
            ws.cell(row=1, column=col).value = None
        # 清除数据行（第2-11行）
        for row in range(2, 12):
            for col in range(2, 5):  # B-D列（对比项保留在A列）
                ws.cell(row=row, column=col).value = None

    # 确定输出路径
    if output_path is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        dir_name = os.path.dirname(template_path)
        output_path = os.path.join(dir_name, f'求职档案_空白版_{timestamp}.xlsx')

    # 保存文件
    wb.save(output_path)
    print(f"已生成干净模板：{output_path}")
    return output_path


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("用法：python clear_template.py <模板路径> [输出路径]")
        sys.exit(1)

    template = sys.argv[1]
    output = sys.argv[2] if len(sys.argv) > 2 else None
    clear_template_data(template, output)
